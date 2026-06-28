import os
import re
import pandas as pd
import openpyxl
import unicodedata

# Dynamic base directory resolution
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

def find_latest_file(base_path, prefix, extension=".xlsx"):
    """
    Finds the latest version file matching the prefix in the base_path or its subdirectories.
    Example: prefix="PU333_01_엔진_v" or ["PU333_엔진_v", "PU333_01_엔진_v"], extension=".xlsx"
    """
    candidates = []
    prefixes = [prefix] if isinstance(prefix, str) else prefix
    if os.path.exists(base_path):
        for root, dirs, files in os.walk(base_path):
            # Exclude cache or test directories
            if any(p in root for p in [".pytest_cache", "__pycache__", ".git", ".gemini"]):
                continue
            for file in files:
                normalized = unicodedata.normalize('NFC', file)
                # Check if file starts with any of the specified prefixes
                starts_with_prefix = any(normalized.startswith(p) for p in prefixes)
                if starts_with_prefix and normalized.endswith(extension):
                    # Extract version number using regex (e.g. v11.1 -> 11.1)
                    match = re.search(r'v(\d+(?:\.\d+)*)', normalized)
                    if match:
                        version_str = match.group(1)
                        try:
                            version_tuple = tuple(map(int, version_str.split('.')))
                            candidates.append((version_tuple, os.path.join(root, file)))
                        except ValueError:
                            pass
    
    if candidates:
        candidates.sort(key=lambda x: x[0], reverse=True)
        return candidates[0][1]
    return None

# Locate the latest DB files
ENGINE_PATH = find_latest_file(BASE_DIR, ["PU333_엔진_v", "PU333_01_엔진_v"])
MANSE_DB_PATH = find_latest_file(BASE_DIR, "PU333_05_만세력DB_v")

# Fallback values if autodetect fails
if not ENGINE_PATH:
    ENGINE_PATH = os.path.join(BASE_DIR, "P333.최신v11.0", "PU333_01_엔진_v11.0.xlsx")
if not MANSE_DB_PATH:
    MANSE_DB_PATH = os.path.join(BASE_DIR, "P333.최신v11.0", "PU333_05_만세력DB_v11.0.xlsx")

print(f"[DB] Autodetected ENGINE_PATH: {ENGINE_PATH}")
print(f"[DB] Autodetected MANSE_DB_PATH: {MANSE_DB_PATH}")

# Cache variables
manse_data = {}  # key: 'YYYY-MM-DD', value: {'목': int, '화': int, '토': int, '금': int, '수': int, '체질': str}
biocode_patterns = {}  # key: '1-1-1', value: {'name': str, 'current': str, 'flaw': str, 'solution': str}
habits_prescriptions = []  # List of habits and tips

def load_databases():
    global manse_data, biocode_patterns, habits_prescriptions
    
    # Reset caches
    manse_data.clear()
    biocode_patterns.clear()
    habits_prescriptions.clear()
    
    # 1. Load Manse DB
    if os.path.exists(MANSE_DB_PATH):
        try:
            df_manse = pd.read_excel(MANSE_DB_PATH, sheet_name='만세력_DB')
            for _, row in df_manse.iterrows():
                date_val = row['날짜']
                if pd.notna(date_val):
                    date_str = pd.to_datetime(date_val).strftime('%Y-%m-%d')
                    manse_data[date_str] = {
                        '목': int(row.get('木', 0)),
                        '화': int(row.get('火', 0)),
                        '토': int(row.get('土', 0)),
                        '금': int(row.get('金', 0)),
                        '수': int(row.get('水', 0)),
                        '체질': "순환-발산형" if "소양인" in str(row.get('최종 체질', '')).strip() else \
                               "수렴-저장형" if "태음인" in str(row.get('최종 체질', '')).strip() else \
                               "민감-조율형" if "소음인" in str(row.get('최종 체질', '')).strip() else \
                               str(row.get('최종 체질', '일반체질')).strip()
                    }
            print(f"[DB] Manse DB loaded: {len(manse_data)} rows cached.")
        except Exception as e:
            print(f"[DB] Error loading Manse DB: {e}")
    else:
        print(f"[DB] Manse DB file not found at: {MANSE_DB_PATH}")

    # 2. Load BioCode 27 Patterns from Engine
    if os.path.exists(ENGINE_PATH):
        try:
            wb = openpyxl.load_workbook(ENGINE_PATH, data_only=True)
            
            if '27코드마스터' in wb.sheetnames:
                sheet = wb['27코드마스터']
                print("[DB] Detecting v13.1 engine layout in '27코드마스터'")
                for r in range(5, 45):
                    code = sheet.cell(r, 1).value
                    if code:
                        code_str = str(code).strip()
                        name_val = str(sheet.cell(r, 5).value or '').strip()
                        body_born = str(sheet.cell(r, 6).value or '').strip()
                        constitution = str(sheet.cell(r, 7).value or '').strip()
                        mind = str(sheet.cell(r, 8).value or '').strip()
                        body_curr = str(sheet.cell(r, 9).value or '').strip()
                        absorption = str(sheet.cell(r, 10).value or '').strip()
                        diet = str(sheet.cell(r, 11).value or '').strip()
                        meal_method = str(sheet.cell(r, 12).value or '').strip()
                        snack = str(sheet.cell(r, 13).value or '').strip()
                        habit = str(sheet.cell(r, 14).value or '').strip()
                        
                        current_desc = f"타고난 체형: {body_born} / 기질 체질: {constitution} (정신성향: {mind})"
                        flaw_desc = f"현재 몸 상태: {body_curr} (대사 흡수 수준: {absorption}) / 극복할 생활 습관: {habit}"
                        solution_desc = f"추천 영양 식단: {diet} ({meal_method}) / 추천 맞춤 간식: {snack}"
                        
                        biocode_patterns[code_str] = {
                            'name': name_val,
                            'current': current_desc,
                            'flaw': flaw_desc,
                            'solution': solution_desc
                        }
            elif '27패턴 마스터(분포+처방)' in wb.sheetnames:
                # v11.1+ layout: B col is Code, G is Name, H is Current, I is Flaw, J is Solution
                sheet = wb['27패턴 마스터(분포+처방)']
                print("[DB] Detecting v11.1+ engine layout in '27패턴 마스터(분포+처방)'")
                for r in range(3, 35):
                    code = sheet.cell(r, 2).value
                    if code:
                        code_str = str(code).strip()
                        biocode_patterns[code_str] = {
                            'name': str(sheet.cell(r, 7).value or '').strip(),
                            'current': str(sheet.cell(r, 8).value or '').strip(),
                            'flaw': str(sheet.cell(r, 9).value or '').strip(),
                            'solution': str(sheet.cell(r, 10).value or '').strip()
                        }
            elif '7. 상담 키포인트 자동 조회' in wb.sheetnames:
                # v11.0- layout: A col is Code, B is Name, C is Current, D is Flaw, E is Solution
                sheet = wb['7. 상담 키포인트 자동 조회']
                print("[DB] Detecting v11.0- engine layout in '7. 상담 키포인트 자동 조회'")
                for r in range(15, 42):
                    code = sheet.cell(r, 1).value
                    if code:
                        code_str = str(code).strip()
                        biocode_patterns[code_str] = {
                            'name': str(sheet.cell(r, 2).value or '').strip(),
                            'current': str(sheet.cell(r, 3).value or '').strip(),
                            'flaw': str(sheet.cell(r, 4).value or '').strip(),
                            'solution': str(sheet.cell(r, 5).value or '').strip()
                        }
            else:
                print("[DB] Warning: No recognized BioCode sheets found in engine workbook.")
            
            # Defense override for 3-3-2 biocode name mismatch (프레임 확장 유망주 -> 성장가능 가치주)
            if '3-3-2' in biocode_patterns:
                if biocode_patterns['3-3-2']['name'] == '프레임 확장 유망주' or not biocode_patterns['3-3-2']['name']:
                    biocode_patterns['3-3-2']['name'] = '성장가능 가치주'
                    print("[DB] Corrected 3-3-2 name to '성장가능 가치주' due to Excel mismatch.")
                
            print(f"[DB] 27 BioCode patterns loaded: {len(biocode_patterns)} patterns cached.")
            
            # Read Habits Prescriptions from '속설·습관 처방' sheet
            if '속설·습관 처방' in wb.sheetnames:
                sheet_habits = wb['속설·습관 처방']
                for r in range(17, 25):
                    habit_name = sheet_habits.cell(r, 2).value
                    if habit_name:
                        habits_prescriptions.append({
                            'habit': str(habit_name).strip(),
                            'status_impact': str(sheet_habits.cell(r, 3).value or '').strip(),
                            'prescription': str(sheet_habits.cell(r, 4).value or '').strip(),
                            'ref': str(sheet_habits.cell(r, 5).value or '').strip()
                        })
                print(f"[DB] Loaded {len(habits_prescriptions)} habits prescriptions.")
            
        except Exception as e:
            print(f"[DB] Error loading Engine DB: {e}")
    else:
        print(f"[DB] Engine DB file not found at: {ENGINE_PATH}")


def get_manse_info(birth_date_str):
    """
    birth_date_str: 'YYYY-MM-DD'
    """
    if birth_date_str in manse_data:
        return manse_data[birth_date_str]
    print(f"[DB] birthdate '{birth_date_str}' not found in Manse DB. Fallback to default.")
    return {
        '목': 1,
        '화': 1,
        '토': 1,
        '금': 1,
        '수': 1,
        '체질': '순환-발산형'
    }

def get_biocode_pattern(code):
    """
    code: '1-2-2', '3-3-3', etc.
    """
    normalized_code = str(code).strip()
    return biocode_patterns.get(normalized_code, {
        'name': '일반 아동 표준',
        'current': '평범한 상태',
        'flaw': '특별히 발견된 취약점 없음',
        'solution': '기본 성장판 자극 운동 및 균형식 섭취'
    })

# Load databases immediately on import
load_databases()
# Force server reload: database updated on 2026-06-25

