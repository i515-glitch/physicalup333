import openpyxl
import os
import unicodedata

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

target_dir_name = 'P333.최신v11.0'
for item in os.listdir(BASE_DIR):
    if unicodedata.normalize('NFC', item) == 'P333.최신v11.0':
        target_dir_name = item
        break

P333_DIR = os.path.join(BASE_DIR, target_dir_name)
engine_file_name = 'PU333_01_엔진_v11.0.xlsx'
for item in os.listdir(P333_DIR):
    if unicodedata.normalize('NFC', item) == 'PU333_01_엔진_v11.0.xlsx':
        engine_file_name = item
        break

ENGINE_PATH = os.path.join(P333_DIR, engine_file_name)
wb = openpyxl.load_workbook(ENGINE_PATH, data_only=True)

print("Comprehensive search for any text containing numbering like '1.' or 'Q' in all sheets:")

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    for r in range(1, 150):
        for c in range(1, 40):
            val = sheet.cell(r, c).value
            if val and isinstance(val, str):
                # Search for numbered lists or questions
                val_clean = val.strip()
                if val_clean.startswith("Q") or any(val_clean.startswith(f"{i}.") for i in range(1, 25)) or "문항" in val_clean or "질문" in val_clean:
                    print(f"[{sheet_name}] R{r}C{c}: {val_clean[:100]}")
